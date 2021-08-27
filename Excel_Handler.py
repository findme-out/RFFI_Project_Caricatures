import os
import pandas as pd
import openpyxl

#----------------------- RFFI Project Caricatures ---------------------
# 2021
# Excel Handler File
#----------------------------------------------------------------------


#обработка аргументов из командный строки: 
#первый аргумент = название файла, второй = адрес до папки с Excel файлами для обработки)
def get_file_location(folder_path):
    # function for the path error handling
    # returns list with all .xls/.xlsx file paths in the folder
    try:
        file_location = []
        path = ""
        for i in folder_path:
            if i == '\\':
                path += '/'
            else:
                path += i
        directory = os.fsencode(path)
        for file in os.listdir(directory):
            filename = os.fsdecode(file)
            if filename.endswith(".xlsx") or filename.endswith(".xls"):
                if path[-1] != '/':
                    path += '/'
                file_location.append(path + filename)
        return file_location
    except BaseException as error:
        print(error)
        quit()

#Парсинг одного файла 
def get_data(file_location, expert_id, proj_id):
    # function for excel file parsing
    # returns list with expert name, cases and their marks
    # returns file parsing status
    file = pd.read_excel(file_location)
    # drop all empty rows and columns
    file = file.dropna(how='all')
    # rename columns for convenience
    file.columns = range(0, len(file.columns))
    status = []
    expert_name = ''
    proj_num_id = []
    proj_num = []
    proj = -1
    # iteration through excel table rows and columns
    for row_index, row_val in file.iterrows():
        temp = []
        proj_status = False
        for col_index, col_val in row_val.items():
            if proj_status:
                temp.append(col_val)
            if col_val == col_val:
                if type(col_val) is str:
                    col_val = col_val.strip().lower()
                    # keyword comparison
                    if col_val == expert_id:
                        expert_name = file[row_index][col_index + 1]
                    elif col_val == proj_id:
                        proj = col_index
                if proj == col_index and type(col_val) is int:
                    proj_status = True
                    proj_num_id.append(col_val)
        if proj_status:
            proj_num.append(temp)
    # format marks
    blank_val = 0
    positive_val = 1
    allowed_chars = ['x', 'X', '1', '+']
    for i in range(len(proj_num)):
        for j in range(len(proj_num[i])):
            if proj_num[i][j] != proj_num[i][j]:
                proj_num[i][j] = blank_val
            elif proj_num[i][j] in allowed_chars:
                proj_num[i][j] = positive_val
            else:
                error_1 = 'Неверный символ оценки'
                if error_1 not in status:
                    status.append(error_1)
    if len(status) != 0:
        return [], status
    status.append('success')
    return [expert_name, proj_num_id, proj_num], status

#запись блоков для отчета
def get_rules(file_location, proj_id):
    file = pd.read_excel(file_location)
    file = file.dropna(how='all')
    file.columns = range(0, len(file.columns))
    proj = -1
    rules = []
    for row_index, row_val in file.iterrows():
        temp_rules = []
        for col_index, col_val in row_val.items():
            if proj == col_index and type(col_val) is int:
                return rules
            if proj != -1 and col_index > proj:
                temp_rules.append(col_val)
            if col_val == col_val:
                if type(col_val) is str:
                    col_val = col_val.strip().lower()
                    if col_val == proj_id:
                        proj = col_index
        if len(temp_rules) != 0:
            rules.append(temp_rules)

#сохранение файла в xlsx-формате
def save_wb(wb, path, file_name):
    # folder creation
    if not os.path.exists(path + 'Отчет'):
        os.mkdir(path + 'Отчет')
    path += 'Отчет/'
    temp = file_name
    index = 1
    # check for existing files with such name in the folder
    while os.path.exists(path + temp + '.xlsx'):
        temp = file_name + '_' + str(index)
        index += 1
    wb.save(path + temp + '.xlsx')

#создание отчета по парсингу файлов
def generate_report(path, file_location, status):
    wb_result = openpyxl.Workbook()
    sheet_result = wb_result.active
    sheet_result.title = 'Отчет'
    x = 1
    y = 1
    for i in range(0, len(file_location)):
        x = 1
        y += 1
        temp = str(file_location[i])
        c = sheet_result.cell(row=y, column=x)
        c.value = temp
        for j in range(0, len(status[i])):
            x += 1
            temp = str(status[i][j])
            c = sheet_result.cell(row=y, column=x)
            c.value = temp
    save_wb(wb_result, path, 'отчет')

#Создание общей таблицы из успешно обработанных файлов (отчет_таблица)
def generate_report_table(path, data, rules, proj_id):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Отчет'
    x = 1
    y = 1
    for i in range(0, len(rules)):
        x = 3
        for j in range(0, len(rules[i])):
            temp = str(rules[i][j])
            if temp != 'nan':
                c = sheet.cell(row=y, column=x)
                c.value = temp
            x += 1
        y += 1
    for i in range(0, len(data['expert_name'])):
        x = 1
        temp = str(data['expert_name'][i])
        c = sheet.cell(row=y, column=x)
        c.value = temp
        x = 2
        y += 1
        temp = str(proj_id)
        c = sheet.cell(row=y, column=x)
        c.value = temp
        for j in range(0, len(data['proj_num_id'][i])):
            x = 2
            y += 1
            temp = data['proj_num_id'][i][j]
            if type(temp) is str:
                temp = str(temp)
            c = sheet.cell(row=y, column=x)
            c.value = temp
            for k in range(0, len(data['proj_num'][i][j])):
                x += 1
                temp = data['proj_num'][i][j][k]
                if type(temp) is str:
                    temp = str(temp)
                c = sheet.cell(row=y, column=x)
                c.value = temp
        y += 1
    save_wb(wb, path, 'отчет_таблица')

# save result dictionary to csv format
def save_csv(data, path, file_name):
    df = pd.DataFrame.from_dict(data)
    # folder creation
    if not os.path.exists(path + 'Отчет'):
        os.mkdir(path + 'Отчет')
    path += 'Отчет/'
    temp = file_name
    index = 1
    # check for existing files with such name in the folder
    while os.path.exists(path + temp + '.csv'):
        temp = file_name + '_' + str(index)
        index += 1
    df.to_csv(path + temp + '.csv', sep=';', encoding='utf-8-sig')

#Формирование результатов обработки (общая матрица для всех экспертов)
def get_result_data(file_location, gen_report=True, gen_report_table=True, gen_report_result=True):
    # dictionary with all parsed data
    # returns dictionary with parsed data and status of files parsing
    data = {
        'expert_name': [],
        'proj_num_id': [],  # № карикатуры
        'proj_num': []  # оценка
    }
    rules = []
    # keywords
    expert_id = 'имя эксперта:'
    proj_id = '№ карикатуры'
    status = []
    file_success = 0
    for location in file_location:
        temp_data, temp_status = get_data(location, expert_id, proj_id)
        status.append(temp_status)
        if temp_status[0] == 'success':
            file_success += 1
            data['expert_name'].append(temp_data[0])
            data['proj_num_id'].append(temp_data[1])
            data['proj_num'].append(temp_data[2])
            if len(rules) == 0:
                rules = get_rules(location, proj_id)
    path = file_location[0][:(file_location[0].rfind('/')) + 1]
    if gen_report:
        generate_report(path, file_location, status)
    if gen_report_table and file_success != 0:
        generate_report_table(path, data, rules, proj_id)
    if gen_report_result and file_success != 0:
        save_csv(data, path, 'отчет_результат')
    return data, status
