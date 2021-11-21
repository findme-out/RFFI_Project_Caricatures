import os
import pandas as pd
import openpyxl

#----------------------- RFFI Project Caricatures ---------------------
# 2021
# Excel Handler File
#----------------------------------------------------------------------


#обработка аргументов из командный строки: 
#первый аргумент = название файла, второй = адрес до папки с Excel файлами для обработки)
def get_file_location(folder_path):
    # function for the path error handling
    # returns list with all .xls/.xlsx file paths in the folder
    try:
        file_location = []
        path = ""
        for i in folder_path:
            if i == '\\':
                path += '/'
            else:
                path += i
        directory = os.fsencode(path)
        for file in os.listdir(directory):
            filename = os.fsdecode(file)
            if filename.endswith(".xlsx") or filename.endswith(".xls"):
                if path[-1] != '/':
                    path += '/'
                file_location.append(path + filename)
        return file_location
    except BaseException as error:
        print(error)
        quit()

#Парсинг одного файла 
def get_data(file_location):
    # function for excel file parsing
    # returns list with cases and their marks
    # returns file parsing status
    file = pd.read_excel(file_location)
    # drop all empty rows and columns
    file = file.dropna(how='all')
    # rename columns for convenience
    file.columns = range(0, len(file.columns))
    status = []
    proj_num_id = []
    proj_num = []
    proj = -1

    def check_int(val):
        try:
            int(val)
            return True
        except ValueError:
            return False


    # iteration through excel table rows and columns
    for row_index, row_val in file.iterrows():
        temp = []
        proj_status = False
        for col_index, col_val in row_val.items():
            if proj_status:
                temp.append(col_val)
            if col_val == col_val:
                if type(col_val) is str:
                    col_val = col_val.strip().lower()
                if proj == -1 and check_int(col_val):
                    proj = col_index
                if proj == col_index and check_int(col_val):
                    proj_status = True
                    proj_num_id.append(col_val)
        if proj_status:
            proj_num.append(temp)
    # format marks
    blank_val = 0
    positive_val = 1
    allowed_chars = ['x', 'X', '1', '+']
    for i in range(len(proj_num)):
        for j in range(len(proj_num[i])):
            if proj_num[i][j] != proj_num[i][j]:
                proj_num[i][j] = blank_val
            elif proj_num[i][j] in allowed_chars:
                proj_num[i][j] = positive_val
            else:
                error_1 = 'Неверный символ оценки'
                if error_1 not in status:
                    status.append(error_1)
    if len(status) != 0:
        return [], status
    status.append('success')
    return [proj_num_id, proj_num], status

#запись блоков для отчета
def get_rules(file_location):
    file = pd.read_excel(file_location)
    file = file.dropna(how='all')
    file.columns = range(0, len(file.columns))
    proj = -1
    end_cond = 4
    rules = []
    for row_index, row_val in file.iterrows():
        temp_rules = []
        for col_index, col_val in row_val.items():
            if proj == -1 and col_val == col_val:
                if type(col_val) is str:
                    proj = col_index
            if proj != -1 and col_index >= proj:
                temp_rules.append(col_val)
        if proj != -1:
            if row_index == end_cond:
                return rules
            rules.append(temp_rules)

# get rules with specific format (only the lowest rule in rule column)
def get_rules_format(file_location):
    file = pd.read_excel(file_location)
    file = file.dropna(how='all')
    file.columns = range(0, len(file.columns))
    rules_format = []
    for i in file:
        last_rule = None
        for j in file[i][:3]:
            if j == j:
                last_rule = j
        if last_rule is not None:
            rules_format.append(last_rule)
    return rules_format


#сохранение файла в xlsx-формате
def save_wb(wb, path, file_name):
    # folder creation
    if not os.path.exists(path + 'Отчет'):
        os.mkdir(path + 'Отчет')
    path += 'Отчет/'
    temp = file_name
    index = 1
    # check for existing files with such name in the folder
    while os.path.exists(path + temp + '.xlsx'):
        temp = file_name + '_' + str(index)
        index += 1
    wb.save(path + temp + '.xlsx')

#создание отчета по парсингу файлов
def generate_report(path, file_location, status):
    wb_result = openpyxl.Workbook()
    sheet_result = wb_result.active
    sheet_result.title = 'Отчет'
    x = 1
    y = 1
    for i in range(0, len(file_location)):
        x = 1
        y += 1
        temp = str(file_location[i])
        c = sheet_result.cell(row=y, column=x)
        c.value = temp
        for j in range(0, len(status[i])):
            x += 1
            temp = str(status[i][j])
            c = sheet_result.cell(row=y, column=x)
            c.value = temp
    save_wb(wb_result, path, 'отчет')

#Создание общей таблицы из успешно обработанных файлов (отчет_таблица)
def generate_report_table(path, data, rules):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Отчет'
    x = 1
    y = 1
    for i in range(0, len(rules)):
        x = 3
        for j in range(0, len(rules[i])):
            temp = str(rules[i][j])
            if temp != 'nan':
                c = sheet.cell(row=y, column=x)
                c.value = temp
            x += 1
        y += 1
    for i in range(0, len(data['proj_num_id'])):
        for j in range(0, len(data['proj_num_id'][i])):
            x = 2
            y += 1
            temp = data['proj_num_id'][i][j]
            if type(temp) is str:
                temp = str(temp)
            c = sheet.cell(row=y, column=x)
            c.value = temp
            for k in range(0, len(data['proj_num'][i][j])):
                x += 1
                temp = data['proj_num'][i][j][k]
                if type(temp) is str:
                    temp = str(temp)
                c = sheet.cell(row=y, column=x)
                c.value = temp
        y += 1
    save_wb(wb, path, 'отчет_таблица')

# save result dictionary to csv format
def save_csv(data, path, file_name):
    df = pd.DataFrame.from_dict(data, dtype='O')
    # folder creation
    if not os.path.exists(path + 'Отчет'):
        os.mkdir(path + 'Отчет')
    path += 'Отчет/'
    temp = file_name
    index = 1
    # check for existing files with such name in the folder
    while os.path.exists(path + temp + '.csv'):
        temp = file_name + '_' + str(index)
        index += 1
    df.to_csv(path + temp + '.csv', sep=';', encoding='utf-8-sig')

# format data for future analysis (rules - columns with marks as their values)
def format_csv(data, file_path):
    new_data = dict()
    rules_format = get_rules_format(file_path)
    for i in range(0, len(data['proj_num_id'])):
        for j in range(0, len(data['proj_num_id'][i])):
            for k in range(0, len(rules_format)):
                if rules_format[k] not in new_data:
                    new_data[rules_format[k]] = [data['proj_num'][i][j][k]]
                else:
                    new_data[rules_format[k]].append(data['proj_num'][i][j][k])
    return new_data

# handle statistics (maybe should be moved to the new project)
def work_with_format_csv():
    file = pd.read_csv(path_to_csv, sep=';')
    file.drop(file.columns[0], axis=1)
    pass


#Формирование результатов обработки (общая матрица для всех экспертов)
def get_result_data(file_location, gen_report=True, gen_report_table=True, gen_report_result=True):
    # dictionary with all parsed data
    # returns dictionary with parsed data and status of files parsing
    data = {
        'proj_num_id': [],  # № карикатуры
        'proj_num': []  # оценка
    }
    rules = []
    status = []
    file_success = 0
    for location in file_location:
        temp_data, temp_status = get_data(location)
        status.append(temp_status)
        if temp_status[0] == 'success':
            file_success += 1
            data['proj_num_id'].append(temp_data[0])
            data['proj_num'].append(temp_data[1])
            if len(rules) == 0:
                rules = get_rules(location)
    path = file_location[0][:(file_location[0].rfind('/')) + 1]
    if gen_report:
        generate_report(path, file_location, status)
    if gen_report_table and file_success != 0:
        generate_report_table(path, data, rules)
    if gen_report_result and file_success != 0:
        new_data = format_csv(data, file_location[0])
        save_csv(new_data, path, 'отчет_результат')
    return data, status
